import base64
import io
import json
import os
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from flask import (
    Flask, jsonify, render_template, request,
    redirect, url_for, session, send_from_directory, send_file
)
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
from zoneinfo import ZoneInfo
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import BarChart, Reference
import cv2

APP_DIR = Path(__file__).parent.resolve()
CONFIG_PATH = APP_DIR / "config.json"      # Class configuration
PHOTOS_ROOT = APP_DIR / "photos"          # we now KEEP a copy for history page
PHOTOS_ROOT.mkdir(exist_ok=True)
DATA_DIR = APP_DIR / "data"               # Directory for class-specific data
DATA_DIR.mkdir(exist_ok=True)

DB_PATH = DATA_DIR / "attendance.db"     # SQLite database for hall passes

def init_db():
    """Initialize the SQLite database with hall pass tracking table"""
    with sqlite3.connect(DB_PATH) as conn:
        with open(APP_DIR / "migrations/create_hall_pass_table.sql") as f:
            conn.executescript(f.read())

# ---- Admin password + session secret
ADMIN_PASSWORD = "abhiMora1!"             # <--- you set this
SECRET_KEY = os.getenv("FLASK_SECRET_KEY", "change-me-please")
# -------------------------------------

app = Flask(__name__)
app.secret_key = SECRET_KEY

# Initialize database
init_db()

# Initialize camera
cap = None

def init_camera():
    """Initialize the camera"""
    global cap
    cap = cv2.VideoCapture(0)
    return cap is not None

def camera_is_initialized():
    """Check if camera is initialized"""
    global cap
    return cap is not None and cap.isOpened()

def capture_photo(photo_path):
    """Capture a photo from the camera"""
    global cap
    if not camera_is_initialized():
        return False

    # Create directory if it doesn't exist
    photo_dir = Path(photo_path).parent
    photo_dir.mkdir(parents=True, exist_ok=True)

    ret, frame = cap.read()
    if ret:
        cv2.imwrite(str(photo_path), frame)
        return True
    return False

try:
    init_camera()
except Exception as e:
    print(f"Failed to initialize camera: {e}")
    print("Running without camera support")


def load_config():
    """Load the class configuration"""
    if not CONFIG_PATH.exists():
        raise FileNotFoundError("config.json not found")
    with open(CONFIG_PATH, 'r') as f:
        return json.load(f)

def get_class_config(class_id=None):
    """Get configuration for a specific class or default class"""
    config = load_config()
    if class_id:
        print(f"Looking for class config with id: {class_id}")
        for class_config in config['classes']:
            if class_config['id'] == class_id:
                print(f"Found class config: {class_config}")
                return class_config
        print(f"No class config found for id: {class_id}, using default")
    # Return default class config if no specific class found
    return next((c for c in config['classes'] if c['id'] == config['defaultClass']), 
                config['classes'][0] if config['classes'] else None)

def get_roster_path(class_config):
    """Get the roster file path for a specific class"""
    path = DATA_DIR / class_config['studentsPath']
    path.parent.mkdir(parents=True, exist_ok=True)  # Ensure parent directory exists
    return path

def load_roster(class_id=None):
    """Load roster for a specific class"""
    class_config = get_class_config(class_id)
    if not class_config:
        raise FileNotFoundError("No class configuration found")
    print(f"Loading roster for class {class_id}, config: {class_config}")
    
    roster_path = get_roster_path(class_config)
    print(f"Roster path: {roster_path}, exists: {roster_path.exists()}")
    if not roster_path.exists():
        # Create empty roster file if it doesn't exist
        print(f"Creating new roster file at {roster_path}")
        df = pd.DataFrame(columns=["Name", "s-number"])
        df.to_excel(roster_path, index=False)
        return df
    
    df = pd.read_excel(roster_path)
    cols = {c.lower().strip(): c for c in df.columns}
    name_col = cols.get("name") or cols.get("student") or "Name"
    snum_col = cols.get("s-number") or cols.get("s number") or "s-number"
    df = df.rename(columns={name_col: "Name", snum_col: "s-number"})
    df["s-number"] = df["s-number"].astype(str).str.strip()
    df["Name"] = df["Name"].astype(str).str.strip()
    return df[["Name", "s-number"]]


def get_today_str():
    # Explicit US Central (handles CST/CDT correctly)
    return datetime.now(ZoneInfo("America/Chicago")).strftime("%Y-%m-%d")


def get_today_xlsx_path(class_id=None):
    class_prefix = f"{class_id}_" if class_id else ""
    path = DATA_DIR / f"attendance_{class_prefix}{get_today_str()}.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)  # Ensure parent directory exists
    return path


def ensure_workbook(path: Path):
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    # Added a 5th column PhotoPath for the history page (Photo is still embedded)
    ws.append(["S-Number", "Name", "Timestamp", "Photo", "PhotoPath"])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 36
    wb.save(path)
    return wb, ws


def already_checked_in(ws, s_number: str):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip() == s_number:
            return True
    return False


def first_name(full_name: str) -> str:
    return (full_name or "").split()[0] if full_name else ""


def save_and_embed_photo(ws, row_idx: int, data_url: str, s_number: str):
    # Decode data URL to PIL image
    header, b64data = data_url.split(",", 1)
    binary = base64.b64decode(b64data)
    pil_img = Image.open(io.BytesIO(binary))

    # Resize to keep Excel light
    max_w, max_h = 240, 180
    pil_img.thumbnail((max_w, max_h))

    # Keep a copy on disk for the History page
    day_dir = PHOTOS_ROOT / get_today_str()
    day_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(ZoneInfo("America/Chicago")).strftime("%H%M%S")
    filename = f"{s_number}_{ts}.png"
    file_path = day_dir / filename
    pil_img.save(file_path, format="PNG")

    # Embed into Excel (column D)
    xl_img = XLImage(str(file_path))
    cell_addr = f"D{row_idx}"
    ws.add_image(xl_img, cell_addr)
    ws.row_dimensions[row_idx].height = 140

    # Return relative web path for later display
    web_path = f"{get_today_str()}/{filename}"  # served via /photos/<path>
    return web_path


def get_active_hall_pass(class_id: str, s_number: str):
    """Get active hall pass for a student if any exists"""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        # Also get student name from roster for the response
        cur.execute("""
            SELECT h.*, 
                   STRFTIME('%Y-%m-%d %H:%M:%S', h.check_out_time) as formatted_check_out_time,
                   STRFTIME('%Y-%m-%d %H:%M:%S', h.check_in_time) as formatted_check_in_time
            FROM hall_passes h
            WHERE h.class_id = ? AND h.s_number = ? AND h.status = 'active'
            ORDER BY h.check_out_time DESC LIMIT 1
        """, (class_id, s_number))
        return cur.fetchone()

def record_hall_pass_checkout(class_id: str, s_number: str, photo_path: str, 
                            reason: str, duration: int):
    """Record a new hall pass checkout"""
    # First get student name from roster
    roster = load_roster(class_id)
    student = roster[roster['s-number'].astype(str).str.strip() == str(s_number).strip()]
    if student.empty:
        raise ValueError("Student not found in roster")
    name = student.iloc[0]['Name']
    
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        # Create new pass
        cur.execute("""
            INSERT INTO hall_passes 
            (class_id, s_number, name, check_out_time, expected_duration, 
             check_out_photo, check_out_reason, status)
            VALUES (?, ?, ?, datetime('now'), ?, ?, ?, 'active')
        """, (class_id, s_number, name, duration, photo_path, reason))
        return cur.lastrowid

def record_hall_pass_checkin(pass_id: int, photo_path: str, notes: str):
    """Record a hall pass check-in"""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        # First get checkout time
        cur.execute("SELECT check_out_time FROM hall_passes WHERE id = ?", (pass_id,))
        row = cur.fetchone()
        if not row:
            raise ValueError("Hall pass not found")
            
        # Calculate duration in minutes between checkout and now
        # Convert both times to US Central time for consistent calculation
        checkout_time = datetime.strptime(row['check_out_time'], '%Y-%m-%d %H:%M:%S')
        checkout_time = checkout_time.replace(tzinfo=ZoneInfo("UTC"))
        checkout_time = checkout_time.astimezone(ZoneInfo("America/Chicago"))
        
        now = datetime.now(ZoneInfo("America/Chicago"))
        actual_duration = int((now - checkout_time).total_seconds() / 60)
        
        cur.execute("""
            UPDATE hall_passes 
            SET check_in_time = datetime('now'),
                check_in_photo = ?,
                check_in_notes = ?,
                actual_duration = ?,
                status = 'completed'
            WHERE id = ?
        """, (photo_path, notes, actual_duration, pass_id))

def update_overdue_passes():
    """Update status of overdue hall passes"""
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("""
            UPDATE hall_passes 
            SET status = 'overdue'
            WHERE status = 'active'
            AND datetime('now') > datetime(check_out_time, '+' || expected_duration || ' minutes')
        """)

def calculate_analytics(roster, available_dates):
    """Calculate attendance analytics across all sessions"""
    total_sessions = len(available_dates)
    total_students = len(roster)
    
    # Track attendance for each student
    student_records = {}
    
    for _, student in roster.iterrows():
        s_num = str(student["s-number"])
        student_records[s_num] = {
            "name": student["Name"],
            "s_number": s_num,
            "present_count": 0,
            "absent_count": 0
        }
    
    # Count attendance across all dates
    for date in available_dates:
        xlsx_path = APP_DIR / f"attendance_{date}.xlsx"
        if not xlsx_path.exists():
            continue
            
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        present_on_date = set()
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            s_num = str(row[0]).strip()
            present_on_date.add(s_num)
        
        # Update counts
        for s_num in student_records:
            if s_num in present_on_date:
                student_records[s_num]["present_count"] += 1
            else:
                student_records[s_num]["absent_count"] += 1
    
    # Calculate attendance rates
    students_list = []
    total_attendance = 0
    
    for s_num, record in student_records.items():
        if total_sessions > 0:
            attendance_rate = (record["present_count"] / total_sessions) * 100
        else:
            attendance_rate = 0
        
        students_list.append({
            "name": record["name"],
            "s_number": record["s_number"],
            "present_count": record["present_count"],
            "absent_count": record["absent_count"],
            "attendance_rate": attendance_rate
        })
        total_attendance += attendance_rate
    
    # Calculate average attendance
    avg_attendance = total_attendance / total_students if total_students > 0 else 0
    
    # Sort by name
    students_list.sort(key=lambda x: x["name"])
    
    return {
        "total_sessions": total_sessions,
        "total_students": total_students,
        "avg_attendance": avg_attendance,
        "students": students_list
    }


@app.route("/config")
def get_config():
    """Return the class configuration to the frontend"""
    try:
        config = load_config()
        return jsonify(config)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/verify_class_password", methods=["POST"])
def verify_class_password():
    """Verify the password for a specific class"""
    payload = request.get_json(silent=True) or {}
    class_id = payload.get("classId")
    password = payload.get("password")

    if not class_id or not password:
        return jsonify({"ok": False, "error": "Missing classId or password"}), 400

    try:
        config = load_config()
        class_config = next((c for c in config["classes"] if c["id"] == class_id), None)
        if class_config and class_config["password"] == password:
            return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({"ok": False}), 403

@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")


@app.route("/checkin", methods=["POST"])
def checkin():
    payload = request.get_json(force=True)
    s_number = str(payload.get("s_number", "")).strip()
    photo_data_url = payload.get("image_data_url")
    class_id = payload.get("classId")
    print(f"Check-in request for class_id: {class_id}, s_number: {s_number}")

    if not s_number:
        return jsonify({"ok": False, "error": "Missing s-number"}), 400
    if not photo_data_url or not photo_data_url.startswith("data:image/"):
        return jsonify({"ok": False, "error": "Missing or invalid photo"}), 400

    # Get class configuration
    class_config = get_class_config(class_id)
    if not class_config:
        return jsonify({"ok": False, "error": "Invalid class configuration"}), 400

    # Lookup roster for specific class
    roster = load_roster(class_id)
    print(f"Looking up s-number: '{s_number}' in roster with s-numbers: {roster['s-number'].tolist()}")
    # Ensure s-number formats match exactly
    roster["s-number"] = roster["s-number"].astype(str).str.strip()
    s_number = str(s_number).strip()
    print(f"Normalized s-number: '{s_number}' vs roster s-numbers: {roster['s-number'].tolist()}")
    match = roster.loc[roster["s-number"] == s_number]
    if match.empty:
        return jsonify({"ok": False, "error": "S-number not found in this class"}), 404

    full_name = match.iloc[0]["Name"]
    fname = first_name(full_name)

    # Prepare workbook with class-specific path
    xlsx_path = DATA_DIR / f"attendance_{class_config['id']}_{get_today_str()}.xlsx"
    wb, ws = ensure_workbook(xlsx_path)

    # Only first check-in counts today
    if already_checked_in(ws, s_number):
        return jsonify({"ok": True, "status": "already", "first_name": fname})

    # Central time timestamp
    now = datetime.now(ZoneInfo("America/Chicago"))
    timestamp = now.strftime("%Y-%m-%d %H:%M:%S %Z")

    # Append row first (Photo + PhotoPath filled after)
    ws.append([s_number, full_name, timestamp, "", ""])
    row_idx = ws.max_row

    # Embed photo, keep file, and write PhotoPath (col E)
    photo_prefix = f"{class_config['photosPrefix']}_{s_number}"
    web_path = save_and_embed_photo(ws, row_idx, photo_data_url, photo_prefix)
    ws.cell(row=row_idx, column=5, value=web_path)

    wb.save(xlsx_path)

    return jsonify({"ok": True, "status": "new", "first_name": fname})


# ------------------ HISTORY (password-gated) ------------------

def is_authed():
    return session.get("authed") is True


@app.route("/history", methods=["GET", "POST"])
def history():
    # Password prompt / validation
    if request.method == "POST":
        pwd = request.form.get("password", "")
        if pwd == ADMIN_PASSWORD:
            session["authed"] = True
            return redirect(url_for("history"))
        return render_template("history_login.html", error="Incorrect password")

    if not is_authed():
        return render_template("history_login.html")

    # Get selected class and date from query params
    class_id = request.args.get("classId")
    selected_date = request.args.get("date", get_today_str())
    
    # Get class configuration
    class_config = get_class_config(class_id)
    if not class_config:
        return jsonify({"error": "Invalid class ID"}), 400

    # Get all available dates for this class
    available_dates = []
    pattern = f"attendance_{class_config['id']}_*.xlsx"
    for file in sorted(DATA_DIR.glob(pattern), reverse=True):
        date_str = file.stem.replace(f"attendance_{class_config['id']}_", "")
        available_dates.append(date_str)
    
    if not available_dates:
        available_dates = [get_today_str()]

    # Load class-specific roster
    roster = load_roster(class_id)
    roster["s-number"] = roster["s-number"].astype(str)

    # Get data for selected date
    xlsx_path = DATA_DIR / f"attendance_{class_config['id']}_{selected_date}.xlsx"
    present = []
    present_ids = set()

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            s_num = str(row[0]).strip()
            present_ids.add(s_num)
            # Safe read for optional PhotoPath column
            photo_path = ""
            if len(row) > 4 and row[4]:
                photo_path = row[4]
            present.append({
                "s_number": s_num,
                "name": row[1],
                "timestamp": row[2],
                "photo_path": photo_path,
            })

    # Absent = roster - present_ids for selected date
    absent_df = roster[~roster["s-number"].isin(present_ids)].copy()
    absent_df = absent_df.sort_values(by="Name")

    # Calculate analytics across all dates
    analytics = calculate_analytics(roster, available_dates)

    # Load all class configurations for the dropdown
    config = load_config()

    return render_template("history.html",
                           present=present,
                           absent=absent_df.to_dict(orient="records"),
                           selected_date=selected_date,
                           today=get_today_str(),
                           available_dates=available_dates,
                           analytics=analytics,
                           classes=config["classes"],
                           current_class=class_config)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("history"))


# Serve saved photos (e.g., /photos/2025-10-28/151579_231405.png)
@app.route("/photos/<path:filename>")
def serve_photo(filename):
    return send_from_directory(PHOTOS_ROOT, filename, as_attachment=False)


@app.route("/verify_password", methods=["POST"])
def verify_password():
    """Simple JSON endpoint to verify admin password for client-side actions.
    The history login still uses the form POST to /history; this endpoint
    only returns a JSON-OK result so the front-end can verify the same
    ADMIN_PASSWORD without reusing the login form.
    """
    payload = request.get_json(silent=True) or {}
    pwd = payload.get("password", "")
    if pwd == ADMIN_PASSWORD:
        return jsonify({"ok": True})
    return jsonify({"ok": False}), 403


# ---------- ROSTER management endpoints (admin only) ----------
@app.route('/roster', methods=['GET'])
def get_roster():
    if not is_authed():
        return jsonify({'ok': False, 'error': 'unauthorized'}), 403
    
    class_id = request.args.get('classId')
    try:
        df = load_roster(class_id)
        return jsonify({'ok': True, 'students': df.to_dict(orient='records')})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/roster/add', methods=['POST'])
def add_roster():
    if not is_authed():
        return jsonify({'ok': False, 'error': 'unauthorized'}), 403

    payload = request.get_json(force=True)
    class_id = payload.get('classId')
    
    # Get class configuration
    class_config = get_class_config(class_id)
    if not class_config:
        return jsonify({'ok': False, 'error': 'Invalid class ID'}), 400
    
    roster_path = get_roster_path(class_config)
        
    # accept single or bulk
    entries = payload.get('entries') or []
    # entries may be a single dict
    if isinstance(entries, dict):
        entries = [entries]

    if not entries:
        # try single name/s_number
        name = payload.get('name')
        s_number = payload.get('s_number')
        if name and s_number:
            entries = [{'Name': name, 's-number': str(s_number).strip()}]

    if not entries:
        return jsonify({'ok': False, 'error': 'no entries provided'}), 400

    # Load existing roster
    try:
        df = pd.read_excel(roster_path)
    except FileNotFoundError:
        # Create new roster file if it doesn't exist
        df = pd.DataFrame(columns=['Name', 's-number'])

    # Normalize column names
    cols = {c.lower().strip(): c for c in df.columns}
    name_col = cols.get('name') or cols.get('student') or 'Name'
    snum_col = cols.get('s-number') or cols.get('s number') or 's-number'
    # Ensure columns exist
    if name_col not in df.columns:
        df[name_col] = ''
    if snum_col not in df.columns:
        df[snum_col] = ''

    added = []
    for ent in entries:
        n = (ent.get('Name') or ent.get('name') or '').strip()
        s = str(ent.get('s-number') or ent.get('s_number') or '').strip()
        if not n or not s:
            continue
        # check for duplicates by s-number
        if ((df[snum_col].astype(str).str.strip()) == s).any():
            continue
        df = df.append({name_col: n, snum_col: s}, ignore_index=True)
        added.append({'Name': n, 's-number': s})

    # Save back to class-specific roster
    df.to_excel(roster_path, index=False)

    return jsonify({'ok': True, 'added': added})


# ---------- EXPORT endpoints (admin only) ----------
def _make_attendance_export(selected_date: str, class_id: str = None):
    # Build a workbook combining roster and attendance for selected_date
    class_config = get_class_config(class_id)
    roster = load_roster(class_id)
    xlsx_path = DATA_DIR / f"attendance_{class_config['id']}_{selected_date}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance_{selected_date}"

    headers = ["S-Number", "Name", "Timestamp", "Present", "PhotoPath"]
    ws.append(headers)

    present_ids = set()
    attendance_map = {}
    if xlsx_path.exists():
        awb = load_workbook(xlsx_path, data_only=True)
        aws = awb.active
        for row in aws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            s_num = str(row[0]).strip()
            present_ids.add(s_num)
            attendance_map[s_num] = {
                'timestamp': row[2] if len(row) > 2 else '',
                'photo_path': row[4] if len(row) > 4 else ''
            }

    red_fill = PatternFill(start_color='FFEFEF', end_color='FFEFEF', fill_type='solid')

    # Split roster so absentees appear first
    absent_list = []
    present_list = []
    for student in roster.to_dict(orient='records'):
        s = str(student.get('s-number', '')).strip()
        name = student.get('Name', '')
        present = 'Yes' if s in present_ids else 'No'
        ts = attendance_map.get(s, {}).get('timestamp', '')
        photo = attendance_map.get(s, {}).get('photo_path', '')
        record = { 's': s, 'name': name, 'present': present, 'ts': ts, 'photo': photo }
        if present == 'No':
            absent_list.append(record)
        else:
            present_list.append(record)

    # set photo column width and default
    ws.column_dimensions['E'].width = 22

    for record in (absent_list + present_list):
        s = record['s']
        name = record['name']
        present = record['present']
        ts = record['ts']
        photo = record['photo']

        # Append row and embed photo into column E sized to the cell
        ws.append([s, name, ts, present, ""])
        row_idx = ws.max_row

        if photo:
            photo_path = PHOTOS_ROOT / photo
            if photo_path.exists():
                try:
                    xl_img = XLImage(str(photo_path))
                    # Resize image to fit into the cell (approx)
                    # column width 22 -> approx 22*7 = 154 px
                    xl_img.width = 154
                    xl_img.height = 90
                    cell_addr = f"E{row_idx}"
                    ws.add_image(xl_img, cell_addr)
                    ws.row_dimensions[row_idx].height = 70
                except Exception:
                    ws.cell(row=row_idx, column=5, value=photo)
            else:
                ws.cell(row=row_idx, column=5, value=photo)

        # Highlight absentees (they are at top already)
        if present == 'No':
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = red_fill

    return wb


@app.route('/export/students')
def export_students():
    if not is_authed():
        return redirect(url_for('history'))
    
    class_id = request.args.get('classId')
    class_config = get_class_config(class_id)
    if not class_config:
        return jsonify({'error': 'Invalid class ID'}), 400
        
    roster_path = get_roster_path(class_config)
    return send_file(roster_path, as_attachment=True, 
                    download_name=f"students_{class_config['id']}.xlsx")


@app.route('/export/attendance')
def export_attendance():
    if not is_authed():
        return redirect(url_for('history'))
    
    class_id = request.args.get('classId')
    selected_date = request.args.get('date', get_today_str())
    
    wb = _make_attendance_export(selected_date, class_id)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    class_suffix = f"_{class_id}" if class_id else ""
    return send_file(bio, as_attachment=True, 
                    download_name=f"attendance{class_suffix}_{selected_date}.xlsx",
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/analytics')
def export_analytics():
    if not is_authed():
        return redirect(url_for('history'))
    # Generate analytics workbook (simple CSV-like sheet + bar chart)
    # Collect dates
    available_dates = []
    for file in sorted(DATA_DIR.glob("attendance_*.xlsx"), reverse=True):
        date_str = file.stem.replace("attendance_", "")
        available_dates.append(date_str)
    roster = load_roster()
    analytics = calculate_analytics(roster, available_dates)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Analytics'
    ws.append(['Name', 'S-Number', 'Present Count', 'Absent Count', 'Attendance Rate'])
    for s in analytics['students']:
        ws.append([s['name'], s['s_number'], s['present_count'], s['absent_count'], s['attendance_rate']])

    # Add a simple bar chart for present_count
    chart = BarChart()
    chart.title = 'Present Count per Student'
    chart.y_axis.title = 'Present Count'
    chart.x_axis.title = 'Student'
    data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, 'H2')

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name='analytics.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---------- HALL PASS endpoints ----------
@app.route('/api/hall-pass/checkout', methods=['POST'])
def hall_pass_checkout():
    """Endpoint to check out a hall pass"""
    data = request.get_json()
    class_id = data.get('class_id')
    s_number = data.get('s_number') 
    reason = data.get('reason')
    duration = data.get('duration', 10)  # Default 10 minutes
    image_data_url = data.get('image_data_url')

    # Validate input
    if not all([class_id, s_number, reason, image_data_url]):
        return jsonify({'error': 'Missing required fields'}), 400

    # Load roster to validate student
    try:
        roster = load_roster(class_id)
        student = roster[roster['s-number'].astype(str).str.strip() == str(s_number).strip()]
        if student.empty:
            return jsonify({'error': 'Student not found in roster'}), 404
        student_name = student.iloc[0]['Name']
    except Exception as e:
        print(f"Error loading roster: {e}")
        return jsonify({'error': 'Error validating student'}), 500

    # Check if student already has active pass
    active_pass = get_active_hall_pass(class_id, s_number)
    if active_pass:
        return jsonify({
            'error': 'Student already has active hall pass',
            'pass': dict(active_pass)
        }), 400

    # Take checkout photo
    photo_path = None
    day_dir = PHOTOS_ROOT / "hall_pass"
    day_dir.mkdir(parents=True, exist_ok=True)
    
    if camera_is_initialized():
        timestamp = datetime.now(ZoneInfo("America/Chicago")).strftime('%Y-%m-%d_%H-%M-%S')
        photo_path = str(PHOTOS_ROOT / "hall_pass" / f"checkout_{s_number}_{timestamp}.jpg")
        if capture_photo(photo_path):
            photo_path = f"hall_pass/checkout_{s_number}_{timestamp}.jpg"

    try:
        # Record checkout
        pass_id = record_hall_pass_checkout(
            class_id, s_number, photo_path, reason, duration
        )

        return jsonify({
            'pass_id': pass_id,
            'photo_path': photo_path,
            'student_name': student_name
        })
    except Exception as e:
        print(f"Error recording hall pass checkout: {e}")
        return jsonify({'error': 'Failed to record hall pass checkout'}), 500

@app.route('/api/hall-pass/checkin', methods=['POST'])
def hall_pass_checkin():
    """Endpoint to check in a hall pass"""
    data = request.get_json()
    pass_id = data.get('pass_id')
    notes = data.get('notes', '')

    if not pass_id:
        return jsonify({'error': 'Missing pass ID'}), 400

    # Get the current pass to get student info for the photo
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM hall_passes WHERE id = ?", (pass_id,))
        hall_pass = cur.fetchone()
        
        if not hall_pass:
            return jsonify({'error': 'Hall pass not found'}), 404
        if hall_pass['status'] != 'active':
            return jsonify({'error': 'Hall pass is not active'}), 400

        s_number = hall_pass['s_number']

    # Take checkin photo 
    photo_path = None
    day_dir = PHOTOS_ROOT / "hall_pass"
    day_dir.mkdir(parents=True, exist_ok=True)

    if camera_is_initialized():
        timestamp = datetime.now(ZoneInfo("America/Chicago")).strftime('%Y-%m-%d_%H-%M-%S')
        photo_path = str(PHOTOS_ROOT / "hall_pass" / f"checkin_{s_number}_{timestamp}.jpg")
        if capture_photo(photo_path):
            photo_path = f"hall_pass/checkin_{s_number}_{timestamp}.jpg"

    try:
        # Record checkin
        record_hall_pass_checkin(pass_id, photo_path, notes)
        return jsonify({
            'status': 'success',
            'photo_path': photo_path
        })
    except Exception as e:
        print(f"Error recording hall pass check-in: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/hall-pass/status/<class_id>')
def hall_pass_status(class_id):
    """Get status of all active hall passes for a class"""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        # Update any overdue passes first
        update_overdue_passes()
        
        # Get active and overdue passes
        cur.execute("""
            SELECT * FROM hall_passes 
            WHERE class_id = ? 
            AND status IN ('active', 'overdue')
            ORDER BY check_out_time DESC
        """, (class_id,))
        passes = [dict(row) for row in cur.fetchall()]
        
        return jsonify(passes)

@app.route('/api/hall-pass/history')
def hall_pass_history():
    """Get hall pass history filtered by class and date"""
    if not is_authed():
        return jsonify({'error': 'Unauthorized'}), 403
        
    class_id = request.args.get('classId')
    date = request.args.get('date')
    
    if not class_id:
        return jsonify({'error': 'Missing class ID'}), 400
        
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        if date:
            # Get passes for specific date
            cur.execute("""
                SELECT *, 
                    strftime('%s', check_in_time) - strftime('%s', check_out_time) as duration_seconds
                FROM hall_passes 
                WHERE class_id = ? 
                AND date(check_out_time) = ?
                ORDER BY check_out_time DESC
            """, (class_id, date))
        else:
            # Get all passes
            cur.execute("""
                SELECT *,
                    strftime('%s', check_in_time) - strftime('%s', check_out_time) as duration_seconds
                FROM hall_passes 
                WHERE class_id = ?
                ORDER BY check_out_time DESC
            """, (class_id,))
            
        passes = []
        for row in cur.fetchall():
            pass_data = dict(row)
            # Calculate duration in minutes for completed passes
            if pass_data['check_in_time']:
                duration_secs = pass_data.get('duration_seconds')
                if duration_secs:
                    pass_data['actual_duration_mins'] = int(int(duration_secs) / 60)
            passes.append(pass_data)
            
        return jsonify(passes)

@app.route('/api/hall-pass/export')
def export_hall_passes():
    """Export hall pass data to Excel"""
    if not is_authed():
        return jsonify({'error': 'Unauthorized'}), 403
        
    class_id = request.args.get('classId')
    date = request.args.get('date')
    
    if not class_id:
        return jsonify({'error': 'Missing class ID'}), 400
        
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hall Pass Log'
    
    # Headers
    headers = ['Student ID', 'Name', 'Check Out Time', 'Check In Time', 
              'Duration (min)', 'Reason', 'Notes', 'Status']
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        if date:
            cur.execute("""
                SELECT *,
                    CASE 
                        WHEN check_in_time IS NOT NULL 
                        THEN actual_duration
                        ELSE NULL 
                    END as duration_mins
                FROM hall_passes 
                WHERE class_id = ? 
                AND date(check_out_time) = ?
                ORDER BY check_out_time DESC
            """, (class_id, date))
        else:
            cur.execute("""
                SELECT *,
                    CASE 
                        WHEN check_in_time IS NOT NULL 
                        THEN actual_duration
                        ELSE NULL 
                    END as duration_mins
                FROM hall_passes 
                WHERE class_id = ?
                ORDER BY check_out_time DESC
            """, (class_id,))
            
        for row in cur.fetchall():
            ws.append([
                row['s_number'],
                row['name'],
                row['check_out_time'],
                row['check_in_time'] or 'Not returned',
                row['duration_mins'] or 'N/A',
                row['check_out_reason'],
                row['check_in_notes'] or '',
                row['status']
            ])
            
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    filename = f'hall_pass_log_{class_id}'
    if date:
        filename += f'_{date}'
    filename += '.xlsx'
    
    return send_file(
        bio,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


if __name__ == "__main__":
    # For Chromebook local testing
    app.run(host="0.0.0.0", port=5006, debug=True)